"""
CMHC mortgage rates ingestion — parses the CMHC Housing Information Monthly
'Mortgage Rates Quoted by Institutional Lenders' Excel file and uploads a
normalized CSV to S3.

The file covers national-level posted rates for 1-year, 3-year, and 5-year
fixed mortgages, monthly. Source: CANNEX Financial Exchanges.

Expected file: data/cmhc/mortgage-rates-*.xlsx  (any edition)

Usage:
    python -m src.ingest.mortgage_rates
    python -m src.ingest.mortgage_rates --file data/cmhc/mortgage-rates-02-26-en.xlsx
"""

import argparse
import csv
import io
import logging
from pathlib import Path

import boto3
import openpyxl
from dotenv import load_dotenv

from src.config import AWS_REGION, S3_BUCKET_RAW

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

s3 = boto3.client("s3", region_name=AWS_REGION)

CMHC_DROP_FOLDER = Path("data/cmhc")
S3_KEY = "cmhc/raw/mortgage_rates.csv"

MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3,  "APR": 4,  "MAY": 5,  "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9,  "OCT": 10, "NOV": 11, "DEC": 12,
}

TERM_COLS = {
    2: "1 year",
    3: "3 year",
    4: "5 year",
}


def parse_file(path: Path) -> list[dict]:
    log.info("Parsing %s", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if "Table K1" not in wb.sheetnames:
        raise ValueError(f"Expected 'Table K1' sheet, found: {wb.sheetnames}")

    ws = wb["Table K1"]
    rows = []
    current_year = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        # Year is only populated on the first month of each year (numeric value)
        if row[0] is not None:
            try:
                current_year = int(row[0])
            except (TypeError, ValueError):
                continue  # footer / source row

        month_str = str(row[1]).strip().upper() if row[1] else None
        if not month_str or month_str not in MONTH_MAP:
            continue

        month = MONTH_MAP[month_str]

        for col_idx, term in TERM_COLS.items():
            raw = row[col_idx] if len(row) > col_idx else None
            if raw is None:
                continue
            try:
                rate = round(float(raw), 4)
            except (TypeError, ValueError):
                continue
            rows.append({
                "year":  current_year,
                "month": month,
                "term":  term,
                "rate":  rate,
            })

    log.info("Extracted %d rows (%d months × 3 terms)", len(rows), len(rows) // 3)
    return rows


def upload(rows: list[dict]) -> None:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["year", "month", "term", "rate"])
    writer.writeheader()
    writer.writerows(rows)
    s3.put_object(
        Bucket=S3_BUCKET_RAW,
        Key=S3_KEY,
        Body=buf.getvalue().encode(),
        ContentType="text/csv",
    )
    log.info("Uploaded → s3://%s/%s", S3_BUCKET_RAW, S3_KEY)


def main(file_path: str | None = None) -> None:
    if file_path:
        path = Path(file_path)
    else:
        candidates = sorted(CMHC_DROP_FOLDER.glob("mortgage-rates-*.xlsx"))
        if not candidates:
            raise FileNotFoundError(
                f"No mortgage-rates-*.xlsx file found in {CMHC_DROP_FOLDER}. "
                "Download from CMHC and place it there."
            )
        path = candidates[-1]  # most recent by filename sort
        log.info("Auto-selected: %s", path.name)

    rows = parse_file(path)
    if not rows:
        log.warning("No data extracted — check file format.")
        return
    upload(rows)
    log.info("Mortgage rates ingestion complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest CMHC mortgage rates Excel file")
    parser.add_argument("--file", help="Path to mortgage-rates Excel file")
    args = parser.parse_args()
    main(args.file)
