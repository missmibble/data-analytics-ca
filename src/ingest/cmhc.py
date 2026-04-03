"""
CMHC Rental Market Survey ingestion — parses annual Excel files and uploads
normalized CSVs to S3.

CMHC publishes separate files for vacancy rates and average rents. Place them
in data/cmhc/ — the script detects which type each file contains from the
sheet title. Files must include the year in their name (e.g. 2023_vacancy.xlsx,
2023_rents.xlsx, or simply 2023.xlsx for a combined file).

Usage:
    python -m src.ingest.cmhc                       # process all files in data/cmhc/
    python -m src.ingest.cmhc --file data/cmhc/2023_vacancy.xlsx
"""

import argparse
import csv
import io
import logging
import re
from pathlib import Path

import boto3
import openpyxl
from dotenv import load_dotenv

from src.config import AWS_REGION, CMA_NAME_MAP, S3_BUCKET_RAW

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

s3 = boto3.client("s3", region_name=AWS_REGION)
CMHC_DROP_FOLDER = Path("data/cmhc")

# Column indices in the CSD sheet for each bedroom type (0-indexed)
# Layout: Province(0), Centre(1), CSD(2), DwellingType(3),
#         Bachelor(4), quality(5), 1BR(6), quality(7), 2BR(8), quality(9),
#         3BR+(10), quality(11), Total(12)
BEDROOM_COLS = {
    4:  "Bachelor",
    6:  "1 Bedroom",
    8:  "2 Bedroom",
    10: "3 Bedroom +",
    12: "Total",
}


def extract_year(path: Path) -> int:
    match = re.search(r"(20\d{2})", path.stem)
    if not match:
        raise ValueError(f"Cannot determine year from filename: {path.name}. Include year in filename.")
    return int(match.group(1))


def _csd_sheet_name(wb: openpyxl.Workbook) -> str:
    """Return the CSD sheet name, handling bilingual variants (e.g. 'CSD - SDR')."""
    for name in wb.sheetnames:
        if name.startswith("CSD"):
            return name
    return None


def detect_format(wb: openpyxl.Workbook) -> str:
    """Return 'csd_vacancy', 'csd_rent', or 'arent_vac_occ'."""
    if "ARent_Vac_Occ" in wb.sheetnames:
        return "arent_vac_occ"
    csd = _csd_sheet_name(wb)
    if csd:
        ws = wb[csd]
        title = ""
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            title += " ".join(str(v) for v in row if v)
        log.info("Sheet title: %s", title[:120])
        if "Vacancy" in title:
            return "csd_vacancy"
        if "Rent" in title or "Average" in title:
            return "csd_rent"
    raise ValueError(f"Unrecognised CMHC file format. Sheets: {wb.sheetnames}")


def parse_csd_sheet(wb: openpyxl.Workbook, year: int, data_type: str) -> list[dict]:
    """
    Extract CMA-level rows from the CSD sheet (vacancy rates format).

    CMA totals: Census Subdivision = 'Total' and Dwelling Type = 'Total'.
    Columns: Province(0), Centre(1), CSD(2), DwellingType(3),
             Bachelor(4), quality(5), 1BR(6), quality(7), 2BR(8),
             quality(9), 3BR+(10), quality(11), Total(12)
    """
    ws = wb[_csd_sheet_name(wb)]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[2] != "Total" or r[3] != "Total":
            continue
        centre = str(r[1]).strip() if r[1] else None
        if not centre or centre == "Total":
            continue
        canonical = CMA_NAME_MAP.get(centre)
        if not canonical:
            continue

        for col_idx, bedroom_type in BEDROOM_COLS.items():
            raw_val = r[col_idx] if len(r) > col_idx else None
            if raw_val in (None, "--", "**", ""):
                continue
            val_str = str(raw_val).strip().replace("%", "").replace("$", "").replace(",", "")
            try:
                value = float(val_str)
            except ValueError:
                continue
            rows.append({
                "year": year,
                "centre": canonical,
                "bedroom_type": bedroom_type,
                "value": value,
                "data_type": data_type,
            })
    return rows


# Occupied Units column indices in the ARent_Vac_Occ sheet for each bedroom type.
# Layout per bedroom: Vacant(n), quality, Occupied(n+2), quality, Y/N
# Starting offsets: Bachelor=2, 1BR=7, 2BR=12, 3BR+=17, Total=22; Occupied = offset+2
ARENT_OCCUPIED_COLS = {
    4:  "Bachelor",
    9:  "1 Bedroom",
    14: "2 Bedroom",
    19: "3 Bedroom +",
    24: "Total",
}


def parse_arent_sheet(wb: openpyxl.Workbook, year: int) -> list[dict]:
    """
    Extract CMA-level occupied-unit average rents from the ARent_Vac_Occ sheet.

    CMA rows: Zone column ends with ' CMA' (excluding 'Remainder of CMA').
    Year is in column 1. Occupied Units values are at fixed column indices.
    """
    ws = wb["ARent_Vac_Occ"]
    rows = []
    for r in ws.iter_rows(min_row=11, values_only=True):
        zone = str(r[0]).strip() if r[0] else None
        if not zone or not zone.endswith("CMA"):
            continue
        if "Remainder" in zone:
            continue
        # Strip " CMA" suffix and look up canonical name
        centre_raw = zone[:-4].strip()  # remove " CMA"
        canonical = CMA_NAME_MAP.get(centre_raw)
        if not canonical:
            continue

        row_year = int(r[1]) if r[1] else year
        for col_idx, bedroom_type in ARENT_OCCUPIED_COLS.items():
            raw_val = r[col_idx] if len(r) > col_idx else None
            if raw_val in (None, "--", "**", ""):
                continue
            try:
                value = float(str(raw_val).replace(",", "").strip())
            except (ValueError, TypeError):
                continue
            rows.append({
                "year": row_year,
                "centre": canonical,
                "bedroom_type": bedroom_type,
                "value": value,
                "data_type": "rent",
            })
    return rows


def process_file(path: Path) -> None:
    year = extract_year(path)
    log.info("Processing %s (year=%d)", path.name, year)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    log.info("Sheets: %s", wb.sheetnames)

    fmt = detect_format(wb)
    log.info("Detected format: %s", fmt)

    if fmt == "arent_vac_occ":
        rows = parse_arent_sheet(wb, year)
        data_type = "rent"
    elif fmt == "csd_vacancy":
        rows = parse_csd_sheet(wb, year, "vacancy")
        data_type = "vacancy"
    elif fmt == "csd_rent":
        rows = parse_csd_sheet(wb, year, "rent")
        data_type = "rent"
    else:
        raise ValueError(f"Unknown format: {fmt}")

    cmas = {r["centre"] for r in rows}
    log.info("Extracted %d rows for %d CMAs: %s", len(rows), len(cmas), sorted(cmas))

    if not rows:
        log.warning("No data rows extracted — check CMA_NAME_MAP and file format.")
        return

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["year", "centre", "bedroom_type", "value", "data_type"])
    writer.writeheader()
    writer.writerows(rows)

    key = f"cmhc/raw/{year}_{data_type}.csv"
    s3.put_object(
        Bucket=S3_BUCKET_RAW,
        Key=key,
        Body=buf.getvalue().encode(),
        ContentType="text/csv",
    )
    log.info("Uploaded → s3://%s/%s", S3_BUCKET_RAW, key)


def main(file_path: str | None = None) -> None:
    if file_path:
        paths = [Path(file_path)]
    else:
        paths = sorted(CMHC_DROP_FOLDER.glob("*.xlsx")) + sorted(CMHC_DROP_FOLDER.glob("*.xls"))

    if not paths:
        log.warning("No Excel files found in %s — place CMHC files there and re-run.", CMHC_DROP_FOLDER)
        return

    errors: list[str] = []
    for path in paths:
        try:
            process_file(path)
        except Exception as exc:
            log.error("Failed to process %s: %s", path.name, exc)
            errors.append(path.name)

    if errors:
        raise SystemExit(f"CMHC ingestion failed for: {', '.join(errors)}")
    log.info("CMHC ingestion complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Parse CMHC Excel files and upload to S3")
    parser.add_argument("--file", help="Path to a single CMHC Excel file")
    args = parser.parse_args()
    main(args.file)
