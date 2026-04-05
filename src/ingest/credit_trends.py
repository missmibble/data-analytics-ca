"""
CMHC National Mortgage and Credit Trends ingestion — parses the quarterly
Excel report and uploads a normalized CSV to S3.

Sheets extracted:
  DATA_1  — Mortgage delinquency rates (Canada, Montreal, Toronto, Vancouver)
  DATA_7  — Delinquency rates by credit type (HELOC, credit card, auto, LOC) — national
  DATA_13 — Average credit scores by mortgage status — national
  DATA_25 — Average monthly mortgage payment (existing vs. new) — national

Quarters map to end-of-quarter month: Q1→3, Q2→6, Q3→9, Q4→12.

Usage:
    python -m src.ingest.credit_trends
    python -m src.ingest.credit_trends --file data/cmhc/mortgage-consumer-credit-trends-canada-2025-q4-en.xlsx
"""

import argparse
import csv
import io
import logging
import re
from pathlib import Path

import boto3
import openpyxl
from dotenv import load_dotenv

from src.config import AWS_REGION, S3_BUCKET_RAW

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

s3 = boto3.client("s3", region_name=AWS_REGION)

CMHC_DROP_FOLDER = Path("data/cmhc")
S3_KEY = "cmhc/raw/credit_trends.csv"

QUARTER_TO_MONTH = {1: 3, 2: 6, 3: 9, 4: 12}


def _parse_period(cell: str, current_year: int | None) -> tuple[int, int] | None:
    """Return (year, quarter) from a period cell like '2021Q4' or 'Q1'."""
    cell = str(cell).strip()
    full = re.match(r"(\d{4})Q(\d)", cell)
    if full:
        return int(full.group(1)), int(full.group(2))
    short = re.match(r"Q(\d)$", cell)
    if short and current_year:
        return current_year, int(short.group(1))
    return None


def _safe_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_data1(ws) -> list[dict]:
    """Mortgage delinquency rates — Canada, Montreal, Toronto, Vancouver."""
    rows = []
    current_year = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        period = _parse_period(row[0], current_year)
        if not period:
            continue
        current_year, quarter = period
        for col_idx, geo in [(1, "Canada"), (2, "Montreal"), (3, "Toronto"), (4, "Vancouver")]:
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            rows.append({
                "year": current_year, "quarter": quarter,
                "indicator_name": "Mortgage delinquency rate",
                "geography": geo, "value": val,
            })
    return rows


def parse_data7(ws) -> list[dict]:
    """Delinquency rates by credit type — national (skip Mortgage; covered by DATA_1 Canada)."""
    indicators = {
        2: "HELOC delinquency rate",
        3: "Credit card delinquency rate",
        4: "Auto loan delinquency rate",
        5: "LOC delinquency rate",
    }
    rows = []
    current_year = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        period = _parse_period(row[0], current_year)
        if not period:
            continue
        current_year, quarter = period
        for col_idx, indicator_name in indicators.items():
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            rows.append({
                "year": current_year, "quarter": quarter,
                "indicator_name": indicator_name,
                "geography": "Canada", "value": val,
            })
    return rows


def parse_data13(ws) -> list[dict]:
    """Average credit scores by mortgage status — national."""
    indicators = {
        1: "Avg credit score - Without mortgage",
        2: "Avg credit score - With mortgage",
        3: "Avg credit score - With new mortgage",
    }
    rows = []
    current_year = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        period = _parse_period(row[0], current_year)
        if not period:
            continue
        current_year, quarter = period
        for col_idx, indicator_name in indicators.items():
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            rows.append({
                "year": current_year, "quarter": quarter,
                "indicator_name": indicator_name,
                "geography": "Canada", "value": val,
            })
    return rows


def parse_data25(ws) -> list[dict]:
    """Average monthly mortgage payment — existing vs. new loans — national, Q4 only."""
    indicators = {
        1: "Avg monthly mortgage payment - Existing",
        2: "Avg monthly mortgage payment - New",
    }
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        period = _parse_period(row[0], None)
        if not period:
            continue
        year, quarter = period
        for col_idx, indicator_name in indicators.items():
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            rows.append({
                "year": year, "quarter": quarter,
                "indicator_name": indicator_name,
                "geography": "Canada", "value": val,
            })
    return rows


def parse_file(path: Path) -> list[dict]:
    log.info("Parsing %s", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    missing = [s for s in ["DATA_1", "DATA_7", "DATA_13", "DATA_25"] if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Expected sheets not found: {missing}. Found: {wb.sheetnames}")

    rows = []
    rows += parse_data1(wb["DATA_1"])
    rows += parse_data7(wb["DATA_7"])
    rows += parse_data13(wb["DATA_13"])
    rows += parse_data25(wb["DATA_25"])

    log.info("Extracted %d rows across 4 sheets", len(rows))
    return rows


def upload(rows: list[dict]) -> None:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["year", "quarter", "indicator_name", "geography", "value"])
    writer.writeheader()
    writer.writerows(rows)
    s3.put_object(
        Bucket=S3_BUCKET_RAW,
        Key=S3_KEY,
        Body=buf.getvalue().encode(),
        ContentType="text/csv",
    )
    log.info("Uploaded → s3://%s/%s", S3_BUCKET_RAW, S3_KEY)


def main(file_path: str | None = None) -> None:
    if file_path:
        path = Path(file_path)
    else:
        candidates = sorted(CMHC_DROP_FOLDER.glob("mortgage-consumer-credit-trends-*.xlsx"))
        if not candidates:
            raise FileNotFoundError(
                f"No mortgage-consumer-credit-trends-*.xlsx found in {CMHC_DROP_FOLDER}."
            )
        path = candidates[-1]
        log.info("Auto-selected: %s", path.name)

    rows = parse_file(path)
    if not rows:
        log.warning("No data extracted — check file format.")
        return
    upload(rows)
    log.info("Credit trends ingestion complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest CMHC credit trends Excel file")
    parser.add_argument("--file", help="Path to mortgage-consumer-credit-trends Excel file")
    args = parser.parse_args()
    main(args.file)
